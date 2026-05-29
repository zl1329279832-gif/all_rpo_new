import os
import shutil
import json
import logging
from datetime import datetime
from typing import List, Optional
from database.db_manager import DatabaseManager
from utils import EXPORT_DIR, BACKUP_DIR

logger = logging.getLogger(__name__)


class ExportBackupService:
    def __init__(self, db: DatabaseManager):
        self.db = db
        os.makedirs(EXPORT_DIR, exist_ok=True)
        os.makedirs(BACKUP_DIR, exist_ok=True)

    def export_delivery_list(self, order_id: int, fmt: str = "xlsx") -> Optional[str]:
        order = self.db.get_order_with_details(order_id)
        if not order:
            return None

        photos = self.db.fetch_all("photos", where="order_id=? AND selected=1",
                                    params=(order_id,), order_by="imported_at ASC")
        delivery_files = self.db.fetch_all("delivery_files", where="order_id=?",
                                            params=(order_id,))
        payments = self.db.fetch_all("payments", where="order_id=?",
                                      params=(order_id,))

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"交付清单_{order.get('order_no', order_id)}_{timestamp}"

        if fmt == "xlsx":
            return self._export_xlsx(filename, order, photos, delivery_files, payments)
        elif fmt == "csv":
            return self._export_csv(filename, order, photos, delivery_files, payments)
        elif fmt == "json":
            return self._export_json(filename, order, photos, delivery_files, payments)
        return None

    def _export_xlsx(self, filename, order, photos, delivery_files, payments) -> Optional[str]:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            logger.error("openpyxl 未安装，无法导出 xlsx")
            return self._export_csv(filename.replace(".xlsx", ""), order, photos, delivery_files, payments)

        wb = Workbook()

        ws = wb.active
        ws.title = "交付清单"
        title_font = Font(size=16, bold=True)
        header_font = Font(size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        ws.merge_cells("A1:F1")
        ws["A1"] = f"交付清单 - {order.get('order_no', '')}"
        ws["A1"].font = title_font

        row = 3
        info = [
            ("客户姓名", order.get("customer_name", "")),
            ("联系电话", order.get("customer_phone", "")),
            ("套餐名称", order.get("package_name", "")),
            ("预约日期", f"{order.get('appointment_date', '')} {order.get('appointment_time', '')}"),
            ("订单金额", f"¥{order.get('amount', 0):.2f}"),
            ("已付金额", f"¥{order.get('paid_amount', 0):.2f}"),
            ("订单状态", order.get("order_status", "")),
            ("付款状态", order.get("payment_status", "")),
        ]
        for label, value in info:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="选片清单").font = Font(size=13, bold=True)
        row += 1
        headers = ["序号", "文件名", "精修状态", "精修备注", "文件大小", "导入时间"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        for i, p in enumerate(photos, 1):
            data = [i, p.original_filename, p.retouch_status, p.retouch_notes,
                    f"{p.file_size / 1024:.1f}KB" if p.file_size else "", p.imported_at or ""]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=str(val))
                cell.border = thin_border
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="付款记录").font = Font(size=13, bold=True)
        row += 1
        pay_headers = ["序号", "付款金额", "付款方式", "付款日期", "备注"]
        for col, h in enumerate(pay_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        for i, p in enumerate(payments, 1):
            data = [i, f"¥{p.amount:.2f}", p.method, p.payment_date, p.notes]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=str(val))
                cell.border = thin_border
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="交付文件").font = Font(size=13, bold=True)
        row += 1
        del_headers = ["序号", "文件名", "文件大小", "交付时间", "备注"]
        for col, h in enumerate(del_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        for i, d in enumerate(delivery_files, 1):
            data = [i, d.file_name, f"{d.file_size / 1024:.1f}KB" if d.file_size else "",
                    d.delivered_at or "", d.notes]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=str(val))
                cell.border = thin_border
            row += 1

        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

        filepath = os.path.join(str(EXPORT_DIR), f"{filename}.xlsx")
        wb.save(filepath)
        return filepath

    def _export_csv(self, filename, order, photos, delivery_files, payments) -> str:
        import csv
        filepath = os.path.join(str(EXPORT_DIR), f"{filename}.csv")
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["交付清单", order.get("order_no", "")])
            writer.writerow(["客户", order.get("customer_name", ""),
                             "电话", order.get("customer_phone", "")])
            writer.writerow(["套餐", order.get("package_name", ""),
                             "金额", f"¥{order.get('amount', 0):.2f}"])
            writer.writerow([])
            writer.writerow(["选片清单"])
            writer.writerow(["序号", "文件名", "精修状态", "备注"])
            for i, p in enumerate(photos, 1):
                writer.writerow([i, p.original_filename, p.retouch_status, p.retouch_notes])
            writer.writerow([])
            writer.writerow(["付款记录"])
            writer.writerow(["序号", "金额", "方式", "日期"])
            for i, p in enumerate(payments, 1):
                writer.writerow([i, f"¥{p.amount:.2f}", p.method, p.payment_date])
        return filepath

    def _export_json(self, filename, order, photos, delivery_files, payments) -> str:
        filepath = os.path.join(str(EXPORT_DIR), f"{filename}.json")
        data = {
            "order": {k: str(v) for k, v in order.items()},
            "selected_photos": [
                {"filename": p.original_filename, "retouch_status": p.retouch_status,
                 "retouch_notes": p.retouch_notes}
                for p in photos
            ],
            "payments": [
                {"amount": p.amount, "method": p.method, "date": p.payment_date, "notes": p.notes}
                for p in payments
            ],
            "delivery_files": [
                {"filename": d.file_name, "size": d.file_size, "notes": d.notes}
                for d in delivery_files
            ],
        }
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return filepath

    def backup_database(self) -> Optional[str]:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"studio_backup_{timestamp}.db"
        backup_path = os.path.join(str(BACKUP_DIR), backup_filename)

        try:
            src = self.db.db_path
            if os.path.exists(src):
                shutil.copy2(src, backup_path)
                logger.info(f"数据库备份成功: {backup_path}")
                return backup_path
            return None
        except Exception as e:
            logger.error(f"备份失败: {e}")
            return None

    def restore_database(self, backup_path: str) -> bool:
        if not os.path.exists(backup_path):
            return False
        try:
            self.db.close()
            shutil.copy2(backup_path, self.db.db_path)
            self.db.connect()
            logger.info(f"数据库恢复成功: {backup_path}")
            return True
        except Exception as e:
            logger.error(f"恢复失败: {e}")
            self.db.connect()
            return False

    def list_backups(self) -> List[str]:
        backups = []
        if os.path.exists(str(BACKUP_DIR)):
            for f in sorted(os.listdir(str(BACKUP_DIR)), reverse=True):
                if f.endswith(".db"):
                    backups.append(os.path.join(str(BACKUP_DIR), f))
        return backups
