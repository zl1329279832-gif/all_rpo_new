import os
import io
import pandas as pd
from typing import Dict, Any, Optional, TYPE_CHECKING
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
import plotly.graph_objects as go
import json

if TYPE_CHECKING:
    from .validation import ValidationReport
    from .alerting import AlertReport


def _create_excel_style(workbook: Workbook) -> Dict[str, Any]:
    styles = {}
    styles["header_font"] = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    styles["title_font"] = Font(name="微软雅黑", size=16, bold=True, color="1976D2")
    styles["normal_font"] = Font(name="微软雅黑", size=10)
    styles["highlight_font"] = Font(name="微软雅黑", size=11, bold=True, color="FF9800")

    styles["header_fill"] = PatternFill(start_color="FF1976D2", end_color="FF1976D2", fill_type="solid")
    styles["title_fill"] = PatternFill(start_color="FFE3F2FD", end_color="FFE3F2FD", fill_type="solid")
    styles["alternate_fill"] = PatternFill(start_color="FFF5F5F5", end_color="FFF5F5F5", fill_type="solid")

    styles["center"] = Alignment(horizontal="center", vertical="center", wrap_text=True)
    styles["left"] = Alignment(horizontal="left", vertical="center", wrap_text=True)
    styles["right"] = Alignment(horizontal="right", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="FFCCCCCC"),
        right=Side(style="thin", color="FFCCCCCC"),
        top=Side(style="thin", color="FFCCCCCC"),
        bottom=Side(style="thin", color="FFCCCCCC"),
    )
    styles["border"] = thin_border

    return styles


def _write_dataframe_to_sheet(
    ws,
    df: pd.DataFrame,
    start_row: int,
    styles: Dict[str, Any],
    title: Optional[str] = None,
) -> int:
    current_row = start_row

    if title:
        ws.cell(row=current_row, column=1, value=title)
        ws.cell(row=current_row, column=1).font = styles["title_font"]
        ws.cell(row=current_row, column=1).fill = styles["title_fill"]
        ws.cell(row=current_row, column=1).alignment = styles["center"]
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=len(df.columns) if not df.empty else 1,
        )
        current_row += 1

    if not df.empty:
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=str(col_name))
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["center"]
            cell.border = styles["border"]
        current_row += 1

        for row_idx, (_, row) in enumerate(df.iterrows()):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = styles["normal_font"]
                cell.alignment = styles["center"] if col_idx > 1 else styles["left"]
                cell.border = styles["border"]
                if row_idx % 2 == 1:
                    cell.fill = styles["alternate_fill"]
            current_row += 1

        for col_idx in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                15, len(str(df.columns[col_idx - 1])) * 2 + 5
            )
    else:
        ws.cell(row=current_row, column=1, value="暂无数据")
        current_row += 1

    current_row += 1
    return current_row


def _write_data_quality_sheet(ws, validation_report: "ValidationReport", styles: Dict[str, Any]) -> None:
    ws.sheet_view.showGridLines = False
    
    current_row = 1
    ws.cell(row=current_row, column=1, value="📊 数据质量摘要")
    ws.cell(row=current_row, column=1).font = Font(name="微软雅黑", size=18, bold=True, color="1976D2")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    current_row += 2

    quality_score = validation_report.quality_score
    score_fill = PatternFill(start_color="FFC8E6C9", end_color="FFC8E6C9", fill_type="solid") if quality_score >= 80 else \
                  PatternFill(start_color="FFFFECB3", end_color="FFFFECB3", fill_type="solid") if quality_score >= 60 else \
                  PatternFill(start_color="FFFFCDD2", end_color="FFFFCDD2", fill_type="solid")
    
    ws.cell(row=current_row, column=1, value="数据质量评分")
    ws.cell(row=current_row, column=1).font = styles["header_font"]
    ws.cell(row=current_row, column=1).fill = styles["header_fill"]
    ws.cell(row=current_row, column=1).alignment = styles["center"]
    ws.cell(row=current_row, column=1).border = styles["border"]
    
    ws.cell(row=current_row, column=2, value=f"{quality_score:.1f} / 100")
    ws.cell(row=current_row, column=2).font = Font(name="微软雅黑", size=14, bold=True, color="1B5E20" if quality_score >= 80 else "F57F17" if quality_score >= 60 else "C62828")
    ws.cell(row=current_row, column=2).fill = score_fill
    ws.cell(row=current_row, column=2).alignment = styles["center"]
    ws.cell(row=current_row, column=2).border = styles["border"]
    current_row += 1

    summary_data = [
        ("问题总数", validation_report.total_issues),
        ("严重问题数", validation_report.critical_issues),
        ("警告问题数", validation_report.warning_issues),
        ("整体有效性", "✅ 有效" if validation_report.overall_valid else "❌ 无效"),
    ]
    
    for label, value in summary_data:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=1).font = styles["normal_font"]
        ws.cell(row=current_row, column=1).border = styles["border"]
        ws.cell(row=current_row, column=1).alignment = styles["left"]
        
        cell = ws.cell(row=current_row, column=2, value=value)
        cell.font = styles["normal_font"]
        cell.border = styles["border"]
        cell.alignment = styles["center"]
        
        if label == "严重问题数" and value > 0:
            cell.fill = PatternFill(start_color="FFFFCDD2", end_color="FFFFCDD2", fill_type="solid")
        elif label == "警告问题数" and value > 0:
            cell.fill = PatternFill(start_color="FFFFECB3", end_color="FFFFECB3", fill_type="solid")
        elif label == "整体有效性" and "有效" in str(value):
            cell.fill = PatternFill(start_color="FFC8E6C9", end_color="FFC8E6C9", fill_type="solid")
        elif label == "整体有效性" and "无效" in str(value):
            cell.fill = PatternFill(start_color="FFFFCDD2", end_color="FFFFCDD2", fill_type="solid")
        
        current_row += 1
    
    current_row += 1
    
    from .validation import get_issues_dataframe
    issues_df = get_issues_dataframe(validation_report)
    current_row = _write_dataframe_to_sheet(ws, issues_df, current_row, styles, "问题清单")
    
    if not issues_df.empty:
        last_row = current_row - 1
        first_data_row = current_row - len(issues_df)
        
        red_fill = PatternFill(start_color="FFFFCDD2", end_color="FFFFCDD2", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFECB3", end_color="FFFFECB3", fill_type="solid")
        blue_fill = PatternFill(start_color="FFB3E5FC", end_color="FFB3E5FC", fill_type="solid")
        
        for row in range(first_data_row, last_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "🔴" in str(cell_value):
                for col in range(1, len(issues_df.columns) + 1):
                    ws.cell(row=row, column=col).fill = red_fill
            elif cell_value and "🟡" in str(cell_value):
                for col in range(1, len(issues_df.columns) + 1):
                    ws.cell(row=row, column=col).fill = yellow_fill
            elif cell_value and "🔵" in str(cell_value):
                for col in range(1, len(issues_df.columns) + 1):
                    ws.cell(row=row, column=col).fill = blue_fill
    
    for col_idx in range(1, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.column_dimensions['E'].width = 50


def _write_alert_sheet(ws, alert_report: "AlertReport", styles: Dict[str, Any]) -> None:
    ws.sheet_view.showGridLines = False
    
    current_row = 1
    ws.cell(row=current_row, column=1, value="⚠️ 预警摘要")
    ws.cell(row=current_row, column=1).font = Font(name="微软雅黑", size=18, bold=True, color="1976D2")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    current_row += 2

    health_score = alert_report.overall_health_score
    score_fill = PatternFill(start_color="FFC8E6C9", end_color="FFC8E6C9", fill_type="solid") if health_score >= 80 else \
                  PatternFill(start_color="FFFFECB3", end_color="FFFFECB3", fill_type="solid") if health_score >= 60 else \
                  PatternFill(start_color="FFFFCDD2", end_color="FFFFCDD2", fill_type="solid")
    
    ws.cell(row=current_row, column=1, value="整体健康评分")
    ws.cell(row=current_row, column=1).font = styles["header_font"]
    ws.cell(row=current_row, column=1).fill = styles["header_fill"]
    ws.cell(row=current_row, column=1).alignment = styles["center"]
    ws.cell(row=current_row, column=1).border = styles["border"]
    
    ws.cell(row=current_row, column=2, value=f"{health_score:.1f} / 100")
    ws.cell(row=current_row, column=2).font = Font(name="微软雅黑", size=14, bold=True, color="1B5E20" if health_score >= 80 else "F57F17" if health_score >= 60 else "C62828")
    ws.cell(row=current_row, column=2).fill = score_fill
    ws.cell(row=current_row, column=2).alignment = styles["center"]
    ws.cell(row=current_row, column=2).border = styles["border"]
    current_row += 1

    summary_data = [
        ("预警总数", alert_report.total_alerts),
        ("🔴 严重预警", alert_report.critical_alerts),
        ("🟡 警告预警", alert_report.warning_alerts),
        ("🔵 提示预警", alert_report.info_alerts),
    ]
    
    for label, value in summary_data:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=1).font = styles["normal_font"]
        ws.cell(row=current_row, column=1).border = styles["border"]
        ws.cell(row=current_row, column=1).alignment = styles["left"]
        
        cell = ws.cell(row=current_row, column=2, value=value)
        cell.font = styles["normal_font"]
        cell.border = styles["border"]
        cell.alignment = styles["center"]
        
        if "严重" in label and value > 0:
            cell.fill = PatternFill(start_color="FFFFCDD2", end_color="FFFFCDD2", fill_type="solid")
        elif "警告" in label and value > 0:
            cell.fill = PatternFill(start_color="FFFFECB3", end_color="FFFFECB3", fill_type="solid")
        
        current_row += 1
    
    current_row += 1
    
    from .alerting import get_alert_summary
    type_summary_df = get_alert_summary(alert_report)
    if not type_summary_df.empty:
        current_row = _write_dataframe_to_sheet(ws, type_summary_df, current_row, styles, "各类型预警数")
    
    current_row += 1
    
    alerts_df = alert_report.to_dataframe()
    current_row = _write_dataframe_to_sheet(ws, alerts_df, current_row, styles, "预警清单")
    
    if not alerts_df.empty:
        last_row = current_row - 1
        first_data_row = current_row - len(alerts_df)
        
        red_fill = PatternFill(start_color="FFFFCDD2", end_color="FFFFCDD2", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFECB3", end_color="FFFFECB3", fill_type="solid")
        blue_fill = PatternFill(start_color="FFB3E5FC", end_color="FFB3E5FC", fill_type="solid")
        
        for row in range(first_data_row, last_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "🔴" in str(cell_value):
                for col in range(1, len(alerts_df.columns) + 1):
                    ws.cell(row=row, column=col).fill = red_fill
            elif cell_value and "🟡" in str(cell_value):
                for col in range(1, len(alerts_df.columns) + 1):
                    ws.cell(row=row, column=col).fill = yellow_fill
            elif cell_value and "🔵" in str(cell_value):
                for col in range(1, len(alerts_df.columns) + 1):
                    ws.cell(row=row, column=col).fill = blue_fill
    
    for col_idx in range(1, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['I'].width = 30


def _write_kpi_sheet(ws, analysis_results: Dict[str, Any], styles: Dict[str, Any]) -> None:
    ws.sheet_view.showGridLines = False
    
    current_row = 1
    ws.cell(row=current_row, column=1, value="🎯 核心指标")
    ws.cell(row=current_row, column=1).font = Font(name="微软雅黑", size=18, bold=True, color="1976D2")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    current_row += 2

    kpi_data = analysis_results.get("kpi_metrics", {})
    
    if kpi_data:
        kpi_items = list(kpi_data.items())
        for i in range(0, len(kpi_items), 3):
            for j in range(3):
                if i + j < len(kpi_items):
                    label, value = kpi_items[i + j]
                    col = j * 2 + 1
                    
                    ws.cell(row=current_row, column=col, value=label)
                    ws.cell(row=current_row, column=col).font = styles["normal_font"]
                    ws.cell(row=current_row, column=col).fill = styles["header_fill"]
                    ws.cell(row=current_row, column=col).font = styles["header_font"]
                    ws.cell(row=current_row, column=col).alignment = styles["center"]
                    ws.cell(row=current_row, column=col).border = styles["border"]
                    ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 1)
                    
                    display_value = f"{value:,.2f}" if isinstance(value, float) else f"{value:,}"
                    if "率" in label or "比" in label:
                        display_value += "%"
                    elif "金额" in label or "营业额" in label or "毛利" in label or "单价" in label:
                        display_value = "¥" + display_value
                    
                    ws.cell(row=current_row + 1, column=col, value=display_value)
                    ws.cell(row=current_row + 1, column=col).font = Font(name="微软雅黑", size=16, bold=True, color="1976D2")
                    ws.cell(row=current_row + 1, column=col).fill = PatternFill(start_color="FFE3F2FD", end_color="FFE3F2FD", fill_type="solid")
                    ws.cell(row=current_row + 1, column=col).alignment = styles["center"]
                    ws.cell(row=current_row + 1, column=col).border = styles["border"]
                    ws.merge_cells(start_row=current_row + 1, start_column=col, end_row=current_row + 1, end_column=col + 1)
            
            current_row += 3
        
        current_row += 1
        
        from .metrics import calculate_period_comparison
        if "merged_df" in analysis_results:
            merged_df = analysis_results["merged_df"]
            try:
                comparison = calculate_period_comparison(merged_df)
                if comparison:
                    comparison_df = pd.DataFrame([
                        {"指标": k, "本期值": v.get("current", 0), "上期值": v.get("previous", 0), "同比变化": f"{v.get('change_percent', 0):+.1f}%"}
                        for k, v in comparison.items()
                    ])
                    current_row = _write_dataframe_to_sheet(ws, comparison_df, current_row, styles, "📈 同比环比分析")
                    
                    if not comparison_df.empty:
                        last_row = current_row - 1
                        first_data_row = current_row - len(comparison_df)
                        for row in range(first_data_row, last_row + 1):
                            cell_value = str(ws.cell(row=row, column=4).value)
                            if "-" in cell_value:
                                ws.cell(row=row, column=4).font = Font(name="微软雅黑", size=10, bold=True, color="C62828")
                            elif "+" in cell_value:
                                ws.cell(row=row, column=4).font = Font(name="微软雅黑", size=10, bold=True, color="2E7D32")
            except Exception:
                pass
    
    for col_idx in range(1, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


def generate_excel_report(
    analysis_results: Dict[str, Any],
    output_path: Optional[str] = None,
    validation_report: Optional["ValidationReport"] = None,
    alert_report: Optional["AlertReport"] = None,
) -> bytes:
    wb = Workbook()
    styles = _create_excel_style(wb)

    ws = wb.active
    ws.title = "报告概览"
    ws.sheet_view.showGridLines = False

    current_row = 1
    ws.cell(row=current_row, column=1, value="连锁餐饮数据分析报告")
    ws.cell(row=current_row, column=1).font = Font(name="微软雅黑", size=20, bold=True, color="1976D2")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    current_row += 2

    ws.cell(row=current_row, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=current_row, column=1).font = styles["normal_font"]
    current_row += 1

    date_range = analysis_results.get("date_range", ("未指定", "未指定"))
    ws.cell(row=current_row, column=1, value=f"分析周期: {date_range[0]} 至 {date_range[1]}")
    ws.cell(row=current_row, column=1).font = styles["normal_font"]
    current_row += 2

    kpi_data = analysis_results.get("kpi_metrics", {})
    if kpi_data:
        ws.cell(row=current_row, column=1, value="核心指标")
        ws.cell(row=current_row, column=1).font = styles["highlight_font"]
        current_row += 1
        kpi_df = pd.DataFrame([{"指标": k, "数值": v} for k, v in kpi_data.items()])
        current_row = _write_dataframe_to_sheet(ws, kpi_df, current_row, styles)

    if validation_report is not None:
        ws_quality = wb.create_sheet("📊 数据质量摘要")
        _write_data_quality_sheet(ws_quality, validation_report, styles)

    if alert_report is not None:
        ws_alert = wb.create_sheet("⚠️ 预警摘要")
        _write_alert_sheet(ws_alert, alert_report, styles)

    if kpi_data:
        ws_kpi = wb.create_sheet("🎯 核心指标")
        _write_kpi_sheet(ws_kpi, analysis_results, styles)

    if "store_metrics" in analysis_results:
        ws_store = wb.create_sheet("门店分析")
        _write_dataframe_to_sheet(ws_store, analysis_results["store_metrics"], 1, styles, "门店经营指标")

    if "dish_metrics" in analysis_results:
        ws_dish = wb.create_sheet("菜品分析")
        _write_dataframe_to_sheet(ws_dish, analysis_results["dish_metrics"].head(30), 1, styles, "菜品销售排行TOP30")

    if "category_metrics" in analysis_results:
        ws_cat = wb.create_sheet("品类分析")
        _write_dataframe_to_sheet(ws_cat, analysis_results["category_metrics"], 1, styles, "品类销售分析")

    if "rfm_result" in analysis_results:
        ws_rfm = wb.create_sheet("会员分析")
        _write_dataframe_to_sheet(ws_rfm, analysis_results["rfm_result"], 1, styles, "会员RFM分析")

    if "promotion_metrics" in analysis_results:
        ws_promo = wb.create_sheet("活动分析")
        _write_dataframe_to_sheet(ws_promo, analysis_results["promotion_metrics"], 1, styles, "促销活动效果分析")

    if "refund_analysis" in analysis_results and "退款原因分布" in analysis_results["refund_analysis"]:
        ws_refund = wb.create_sheet("退款分析")
        refund_df = analysis_results["refund_analysis"]["退款原因分布"]
        _write_dataframe_to_sheet(ws_refund, refund_df, 1, styles, "退款原因分析")

    if "anomaly_stores" in analysis_results:
        ws_anomaly = wb.create_sheet("异常检测")
        _write_dataframe_to_sheet(ws_anomaly, analysis_results["anomaly_stores"], 1, styles, "门店异常检测")

    if "hourly_trend" in analysis_results:
        ws_hourly = wb.create_sheet("时段分析")
        _write_dataframe_to_sheet(ws_hourly, analysis_results["hourly_trend"], 1, styles, "营业时段分析")

    if "dish_combinations" in analysis_results:
        ws_combo = wb.create_sheet("菜品组合")
        _write_dataframe_to_sheet(ws_combo, analysis_results["dish_combinations"].head(30), 1, styles, "热门菜品组合")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    if output_path:
        with open(output_path, "wb") as f:
            f.write(output.getvalue())
        return output_path

    return output.getvalue()


def generate_html_report(
    analysis_results: Dict[str, Any],
    template_path: Optional[str] = None,
    validation_report: Optional["ValidationReport"] = None,
    alert_report: Optional["AlertReport"] = None,
) -> str:
    date_range = analysis_results.get("date_range", ("未指定", "未指定"))
    kpi_data = analysis_results.get("kpi_metrics", {})

    html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>连锁餐饮数据分析报告</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Microsoft YaHei', '微软雅黑', Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 40px 20px;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #1976D2 0%, #0D47A1 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}
        .header h1 {{ font-size: 32px; margin-bottom: 10px; }}
        .header p {{ opacity: 0.9; font-size: 14px; }}
        .content {{ padding: 40px; }}
        .section {{ margin-bottom: 40px; }}
        .section-title {{
            font-size: 24px;
            color: #1976D2;
            border-bottom: 3px solid #FF9800;
            padding-bottom: 10px;
            margin-bottom: 20px;
            display: inline-block;
        }}
        .kpi-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .kpi-card {{
            background: linear-gradient(135deg, #E3F2FD 0%, #BBDEFB 100%);
            padding: 25px;
            border-radius: 12px;
            border-left: 5px solid #1976D2;
        }}
        .kpi-card .label {{ font-size: 14px; color: #666; margin-bottom: 8px; }}
        .kpi-card .value {{ font-size: 28px; font-weight: bold; color: #1976D2; }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        th {{
            background: #1976D2;
            color: white;
            padding: 12px 15px;
            text-align: left;
            font-weight: 600;
        }}
        td {{ padding: 12px 15px; border-bottom: 1px solid #eee; }}
        tr:hover {{ background: #F5F5F5; }}
        tr:nth-child(even) {{ background: #FAFAFA; }}
        tr:nth-child(even):hover {{ background: #F5F5F5; }}
        .footer {{
            background: #F5F5F5;
            padding: 20px;
            text-align: center;
            color: #999;
            font-size: 12px;
        }}
        .badge {{
            display: inline-block;
            padding: 4px 10px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
        }}
        .badge-success {{ background: #C8E6C9; color: #2E7D32; }}
        .badge-warning {{ background: #FFECB3; color: #F57F17; }}
        .badge-danger {{ background: #FFCDD2; color: #C62828; }}
        .badge-info {{ background: #B3E5FC; color: #01579B; }}
        .quality-card {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            border-radius: 15px;
            margin-bottom: 30px;
            display: flex;
            align-items: center;
            justify-content: space-between;
            box-shadow: 0 10px 30px rgba(102, 126, 234, 0.3);
        }}
        .quality-score {{
            text-align: center;
        }}
        .quality-score .number {{
            font-size: 48px;
            font-weight: bold;
        }}
        .quality-score .label {{
            font-size: 14px;
            opacity: 0.9;
        }}
        .quality-details {{
            display: flex;
            gap: 30px;
        }}
        .quality-detail {{
            text-align: center;
        }}
        .quality-detail .value {{
            font-size: 24px;
            font-weight: bold;
        }}
        .quality-detail .label {{
            font-size: 12px;
            opacity: 0.8;
        }}
        .alert-section {{
            background: #FFF8E1;
            border-left: 5px solid #FF9800;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 30px;
        }}
        .alert-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 15px;
        }}
        .alert-header h3 {{
            color: #E65100;
            font-size: 18px;
            display: flex;
            align-items: center;
            gap: 10px;
        }}
        .alert-stats {{
            display: flex;
            gap: 20px;
        }}
        .alert-stat {{
            background: white;
            padding: 10px 20px;
            border-radius: 8px;
            text-align: center;
            min-width: 80px;
        }}
        .alert-stat .count {{
            font-size: 24px;
            font-weight: bold;
        }}
        .alert-stat .label {{
            font-size: 12px;
            color: #666;
        }}
        .alert-critical .count {{ color: #C62828; }}
        .alert-warning .count {{ color: #F57F17; }}
        .alert-info .count {{ color: #01579B; }}
        .kpi-summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .kpi-summary-card {{
            background: linear-gradient(135deg, #E3F2FD 0%, #BBDEFB 100%);
            padding: 25px;
            border-radius: 12px;
            border-left: 5px solid #1976D2;
            position: relative;
            overflow: hidden;
        }}
        .kpi-summary-card::before {{
            content: '';
            position: absolute;
            top: -50%;
            right: -50%;
            width: 100%;
            height: 100%;
            background: radial-gradient(circle, rgba(25,118,210,0.1) 0%, transparent 70%);
        }}
        .kpi-summary-card .label {{
            font-size: 14px;
            color: #666;
            margin-bottom: 8px;
            position: relative;
        }}
        .kpi-summary-card .value {{
            font-size: 32px;
            font-weight: bold;
            color: #1976D2;
            position: relative;
        }}
        .kpi-summary-card .change {{
            font-size: 13px;
            margin-top: 8px;
            position: relative;
        }}
        .change-positive {{ color: #2E7D32; }}
        .change-negative {{ color: #C62828; }}
        .health-score-good {{ background: linear-gradient(135deg, #4CAF50 0%, #2E7D32 100%); }}
        .health-score-medium {{ background: linear-gradient(135deg, #FF9800 0%, #E65100 100%); }}
        .health-score-poor {{ background: linear-gradient(135deg, #f44336 0%, #C62828 100%); }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🍜 连锁餐饮数据分析报告</h1>
            <p>分析周期: {date_range[0]} 至 {date_range[1]}</p>
            <p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
        <div class="content">
"""

    if validation_report is not None:
        quality_score = validation_report.quality_score
        quality_class = "health-score-good" if quality_score >= 80 else "health-score-medium" if quality_score >= 60 else "health-score-poor"
        
        html_content += f"""
            <div class="quality-card {quality_class}">
                <div class="quality-score">
                    <div class="number">{quality_score:.1f}</div>
                    <div class="label">📊 数据质量评分</div>
                </div>
                <div class="quality-details">
                    <div class="quality-detail">
                        <div class="value">{validation_report.total_issues}</div>
                        <div class="label">问题总数</div>
                    </div>
                    <div class="quality-detail">
                        <div class="value">{validation_report.critical_issues}</div>
                        <div class="label">🔴 严重问题</div>
                    </div>
                    <div class="quality-detail">
                        <div class="value">{validation_report.warning_issues}</div>
                        <div class="label">🟡 警告问题</div>
                    </div>
                    <div class="quality-detail">
                        <div class="value">{"✅" if validation_report.overall_valid else "❌"}</div>
                        <div class="label">数据有效性</div>
                    </div>
                </div>
            </div>
"""

    if alert_report is not None:
        html_content += f"""
            <div class="alert-section">
                <div class="alert-header">
                    <h3>⚠️ 预警信息摘要</h3>
                    <div style="display: flex; align-items: center; gap: 15px;">
                        <div style="text-align: center;">
                            <div style="font-size: 28px; font-weight: bold; color: white; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 8px 20px; border-radius: 10px;">
                                {alert_report.overall_health_score:.1f}
                            </div>
                            <div style="font-size: 12px; color: #666; margin-top: 4px;">整体健康分</div>
                        </div>
                    </div>
                </div>
                <div class="alert-stats">
                    <div class="alert-stat alert-critical">
                        <div class="count">{alert_report.critical_alerts}</div>
                        <div class="label">🔴 严重预警</div>
                    </div>
                    <div class="alert-stat alert-warning">
                        <div class="count">{alert_report.warning_alerts}</div>
                        <div class="label">🟡 警告预警</div>
                    </div>
                    <div class="alert-stat alert-info">
                        <div class="count">{alert_report.info_alerts}</div>
                        <div class="label">🔵 提示预警</div>
                    </div>
                    <div class="alert-stat">
                        <div class="count">{alert_report.total_alerts}</div>
                        <div class="label">预警总数</div>
                    </div>
                </div>
            </div>
"""

    if kpi_data:
        try:
            from .metrics import calculate_period_comparison
            comparison = None
            if "merged_df" in analysis_results:
                comparison = calculate_period_comparison(analysis_results["merged_df"])
            
            html_content += """
            <div class="section">
                <h2 class="section-title">🎯 核心指标摘要</h2>
                <div class="kpi-summary">
"""
            kpi_items = list(kpi_data.items())[:8]
            for i, (label, value) in enumerate(kpi_items):
                display_value = f"{value:,.2f}" if isinstance(value, float) else f"{value:,}"
                if "率" in label or "比" in label:
                    display_value += "%"
                elif "金额" in label or "营业额" in label or "毛利" in label or "单价" in label:
                    display_value = "¥" + display_value
                
                change_html = ""
                if comparison:
                    comp_key = None
                    for key in comparison.keys():
                        if key in label or label in key:
                            comp_key = key
                            break
                    if comp_key:
                        change_pct = comparison[comp_key].get("change_percent", 0)
                        change_class = "change-positive" if change_pct >= 0 else "change-negative"
                        change_icon = "↑" if change_pct >= 0 else "↓"
                        change_html = f'<div class="change {change_class}">{change_icon} {change_pct:+.1f}% 较上期</div>'
                
                html_content += f"""
                    <div class="kpi-summary-card">
                        <div class="label">{label}</div>
                        <div class="value">{display_value}</div>
                        {change_html}
                    </div>
"""
            html_content += """
                </div>
            </div>
"""
        except Exception:
            pass

    if kpi_data:
        html_content += """
            <div class="section">
                <h2 class="section-title">📊 核心指标概览</h2>
                <div class="kpi-grid">
"""
        for label, value in kpi_data.items():
            if isinstance(value, float):
                display_value = f"{value:,.2f}"
            else:
                display_value = f"{value:,}"
            if "率" in label or "比" in label:
                display_value += "%"
            elif "金额" in label or "营业额" in label or "毛利" in label or "单价" in label:
                display_value = "¥" + display_value
            html_content += f"""
                    <div class="kpi-card">
                        <div class="label">{label}</div>
                        <div class="value">{display_value}</div>
                    </div>
"""
        html_content += """
                </div>
            </div>
"""

    if "store_metrics" in analysis_results:
        df = analysis_results["store_metrics"].head(10)
        html_content += _df_to_html(df, "🏪 门店经营分析")

    if "dish_metrics" in analysis_results:
        df = analysis_results["dish_metrics"].head(15)
        html_content += _df_to_html(df, "🍲 菜品销售排行TOP15")

    if "category_metrics" in analysis_results:
        df = analysis_results["category_metrics"]
        html_content += _df_to_html(df, "🥡 品类销售分析")

    if "rfm_result" in analysis_results:
        df = analysis_results["rfm_result"].head(20)
        html_content += _df_to_html(df, "👥 会员价值分析")

    if "promotion_metrics" in analysis_results:
        df = analysis_results["promotion_metrics"]
        html_content += _df_to_html(df, "🎁 促销活动效果")

    if "anomaly_stores" in analysis_results:
        df = analysis_results["anomaly_stores"]
        html_content += _df_to_html(df, "⚠️ 门店异常检测")

    if "refund_analysis" in analysis_results and "退款原因分布" in analysis_results["refund_analysis"]:
        df = analysis_results["refund_analysis"]["退款原因分布"]
        html_content += _df_to_html(df, "💸 退款原因分析")

    if "dish_combinations" in analysis_results:
        df = analysis_results["dish_combinations"].head(15)
        html_content += _df_to_html(df, "🍱 热门菜品组合")

    html_content += f"""
        </div>
        <div class="footer">
            <p>本报告由连锁餐饮数据分析平台自动生成 | 数据更新时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>
"""

    return html_content


def _df_to_html(df: pd.DataFrame, title: str) -> str:
    if df.empty:
        return ""

    html = f"""
            <div class="section">
                <h2 class="section-title">{title}</h2>
                <table>
                    <thead>
                        <tr>
"""
    for col in df.columns:
        html += f"<th>{col}</th>"
    html += """
                        </tr>
                    </thead>
                    <tbody>
"""
    for _, row in df.iterrows():
        html += "<tr>"
        for col in df.columns:
            value = row[col]
            if "状态" in str(col):
                if "异常" in str(value):
                    badge_class = "badge-danger"
                elif "正常" in str(value):
                    badge_class = "badge-success"
                else:
                    badge_class = "badge-info"
                html += f'<td><span class="badge {badge_class}">{value}</span></td>'
            elif isinstance(value, float):
                html += f"<td>{value:,.2f}</td>"
            else:
                html += f"<td>{value}</td>"
        html += "</tr>"
    html += """
                    </tbody>
                </table>
            </div>
"""
    return html
